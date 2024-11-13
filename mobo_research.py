#Uses EBAY api to collect info on current & historical motherboard sales
from datetime import datetime, timedelta
import csv 
import json
import os.path
import copy
import math
import sys
import openpyxl
import openpyxl.workbook

import ebayAPIFuncLib

MANUFACTUER_DB = ["Gigabyte", "MSI", "Asus", "Asrock", "ProArt"]

COLUMN_LABELS = ["CHECKED", "BUYABLE", "EST. RETURN", "EST. RETURN %" , "BRAND", "NAME", "LISTING TITLE", "LISTING TYPE", "PRICE", "CUR AVG USED PRICE", "CUR MIN USED PRICE", "CUR MAX USED PRICE", "LISTING LINK"]

GENERIC_MOBO_SEARCH_PARAMS = {
    "q": { #keyswords concatenated into a string => 
        "AND": ["Motherboard"], #"k1 k2" space separated means k1 AND k2
        "OR": ["Z690", "Z790", "B660", "B760", "X670", "X670E", "X870", "X870E", "B850", "B650E"] #"(k1, k2, ...)" comma separated means k1 OR k2 or ... 
                    #could possibly get fancy with OR and AND by having OR be conditionally placed between AND search parameters so we can search for (X or Y) and (A or B)
                    # => "(x, y) (a, b)" is how it would look in url I think
                    #current structing of URL with OR params will just be passed all at once so this conceptual functionality isn't possibly yet. 
        },              
    "filter": { #different filters with each their own formatting => {filter1: params}, {filter2: params}
        "conditionIds": "{7000}", #for parts condition 
        # "itemEndDate": "[..2018-12-14T07:47:48Z]", #item listing end from now till 24 hours 
        "buyingOptions" : "{FIXED_PRICE|AUCTION}",
        }, 
            
    # "sort": "", #sort=price from lowest to highest price (maybe behave strangely with bidding items which I need to look)
    "limit": "30", #number of items to return per page (max 200, default 50)
    "offset": "0" #specifies number of items to skip in result set (control pagination of the output) 
}

USED_MOBO_SALES_SEARCH_PARAMS = {
        "q": { #keyswords concatenated into a string => 
        "AND": ["Motherboard"], #"k1 k2" space separated means k1 AND k2
        "OR": [] #"(k1, k2, ...)" comma separated means k1 OR k2 or ... 
                    #could possibly get fancy with OR and AND by having OR be conditionally placed between AND search parameters so we can search for (X or Y) and (A or B)
                    # => "(x, y) (a, b)" is how it would look in url I think
                    #current structing of URL with OR params will just be passed all at once so this conceptual functionality isn't possibly yet. 
        },              
    "filter": { #different filters with each their own formatting => {filter1: params}, {filter2: params}
        "conditions": "{USED}", #for parts condition 
        "conditionIds": "{2000|2500|3000}" #2000 (certified refurbished), 2500 (seller refurbished), 3000 (used)
        # "itemEndDate": "[..2018-12-14T07:47:48Z]", #item listing end from now till 24 hours 
        }, 
            
    # "sort": "", #sort=price from lowest to highest price (maybe behave strangely with bidding items which I need to look)
    "limit": "10", #number of items to return per page (max 200, default 50)
    # "offset": "0" #specifies number of items to skip in result set (control pagination of the output) 
}

def compileMotherboardData(output_excel_file_name: str):
#Search a generic list of motherboards for sale 
    #Z690, Z790, AMD motherboards..... LGA Sockets only
    #Filter for "for parts" condition

    # GENERIC_MOBO_SEARCH_PARAMS["filter"]["itemEndDate"] = "[.." + (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M:%SZ") + "]"
    response_status, broad_search_results = ebayAPIFuncLib.getSearchResults(GENERIC_MOBO_SEARCH_PARAMS, log_url= True, log_output= True, log_output_path= "broad_search_log_response_debug.json")
    if (response_status != 200):
        raise Exception("Ebay API browse method unsuccessful\nResponse code: " + str(response_status))


    #(TODO) Filter for anything (auctions) that expires within 24 hours from now (newer auctions tend to have lower and therefore less relevant prices)


# (DEPRECATED) #Among list of motherboards found, search with historical data on used (not for parts) motherboards. 
    #Time range should be 6 months or so...
    # motherboard_used_history_search_results = findUsedMotherboardHistory(broad_search_results.keys())
    
#Formally define the Brand and MPN for each motherboard
    print("\nGathering item details...\n")
    modified_broad_search_results = copy.deepcopy(broad_search_results)
    for i in range(0, len(broad_search_results["itemSummaries"])):
        mobo_listing = broad_search_results["itemSummaries"][i]
        mobo_listing["brand"], mobo_listing["mpn"], mobo_listing["itemWebUrl"] = getMotherboardDetailsByItemID(mobo_listing["itemId"], log_no_mpn_api_response= True, ignore_no_mpn_exception= True)

        modified_broad_search_results["itemSummaries"][i] = mobo_listing #include added brand and MPN data into item search summary for item "i" 

        
#Among list of motherboards found, search current pricing of used/refurbished versions of that motherboard
    unique_motherboard_mpn_set = set([listing["mpn"] for listing in modified_broad_search_results["itemSummaries"]])
    motherboard_used_pricing_results_by_mpn = getUsedMotherboardPrices(unique_motherboard_mpn_set, results_limit_per_mpn = 10) 

# #Build a basic spreadsheet with the average historical sale price of each motherboard and the current going price for "for parts" motherboards 
    buildMotherboardSpreadsheet(modified_broad_search_results, motherboard_used_pricing_results_by_mpn, output_excel_sheet_name=output_excel_file_name)
    # with open(output_csv_file_name, "w", newline='') as output_csvfile:
        # writer = csv.writer(output_csvfile)
        # writer.writerows(data)

    return


def getMotherboardDetailsByItemID(item_id: str, log_no_mpn_api_response = True, ignore_no_mpn_exception = False) -> tuple[str, str, str]:
    '''
    Returns a tuple
        0. Brand name 
        1. MPN (manufacturer's product name) 
        2. url (direct link to listing which tends to be shorter than item summaries url)
    '''
    output_dict = {}
    brand = ""
    mpn = ""
    url = ""

    response_status, response = ebayAPIFuncLib.getItemDetails(item_id, log_url=False)
    if (response_status != 200):
        raise Exception("Ebay API item details method unsuccessful\nResponse code: " + str(response_status))
    else:
        if ("brand" in response):
                brand = response["brand"]
        if ("mpn" in response):
            mpn = response["mpn"]
        elif ("localizedAspects" in response): #check localized aspects to get MPN if MPN wasn't specified 
            for aspect_dict in response["localizedAspects"]:
                if (aspect_dict["name"] == "Model"):
                    mpn = aspect_dict["value"]
                    break
        if ("itemWebUrl" in response):
            url = response["itemWebUrl"]
        
        if (mpn == "" and log_no_mpn_api_response == True):
            f = open("error_item_details_response_debug.json", 'w')
            json.dump(response, f, indent = 4)
            if (ignore_no_mpn_exception == False):
                raise Exception("MPN Not found from item details search. Copying return details to file 'error_item_details_debug_response.json'\nItemId = " + item_id)

    return brand, mpn, url



def getUsedMotherboardPrices(motherboard_mpn_set: list, results_limit_per_mpn = 20)-> dict:
    '''
    Replacement for "findUsedMotherboardHistory()"
        (atleast until Marketplace Insights API can be enabled for this project)

    Returns a dictionary where each key pertains to a unique motherboard MPN, and the value is a dictionary describing the min_price, max_price, and avg_price selling price for used/refurbished version of that motherboard currently listed on marketplace. 
    '''
    if ("" in motherboard_mpn_set):
        motherboard_mpn_set.remove("")
        #ignore the empty mpn if present

#Conduct a search for each unique motherboard based on keywords 
    used_motherboard_current_sale_pricing = {}
    for mobo_mpn in motherboard_mpn_set:
 
        #Run a search result on recent sales of the given motherboard
        motherboard_search_params = copy.deepcopy(USED_MOBO_SALES_SEARCH_PARAMS)
        motherboard_search_params["q"]["AND"].append(mobo_mpn)
        motherboard_search_params["limit"] = results_limit_per_mpn
        
        response_status, search_results = ebayAPIFuncLib.getSearchResults(motherboard_search_params, log_url= True)
        if (response_status != 200):
            raise Exception("Ebay API browse method unsuccessful\nResponse code: " + str(response_status))
            return

        #Calculate relevant pricing from recent sales      
        min_price, max_price, avg_price = ebayAPIFuncLib.util_CompileSalePriceStatsOfSearchResults(search_results["itemSummaries"]) if "itemSummaries" in search_results else (0,0,0)

        #Categorize the following motherboard historical pricing data 
        used_motherboard_current_sale_pricing[mobo_mpn] = {"min_price": min_price,
                                                            "max_price": max_price,
                                                            "avg_price": avg_price} 

    return used_motherboard_current_sale_pricing 


 
def findUsedMotherboardHistory(motherboard_mpn_set: list, months_to_search = 6)-> dict:
    '''
    (DEPRECATED)
        - Deprecated because can't use item_sales (Marketplace Insights) API without special allowances from EBay
        Therefore, it's impossible to use api to make historical search until further notice.
        - For replacement, see "getUsedMotherboardPrices()" 

    Each key is paired with a value representing the average historical sale of that motherboard in the used market.

    Returns a dictionary where each key pertains to a unique motherboard, and the value is the average historical sold price of that motherboard within the time frame. 
    '''
#Conduct a historical search for each unique motherboard based on keywords 
    used_motherboard_historical_sale_pricing = {}
    for mobo_mpn in motherboard_mpn_set:
        #Run a search result on recent sales of the given motherboard
        response_status, historical_search_results = ebayAPIFuncLib.getSearchResults(USED_MOBO_SALES_SEARCH_PARAMS, log_url= True)
        if (response_status != 200):
            raise Exception("Ebay API browse method unsuccessful\nResponse code: " + str(response_status))
            return

        #Calculate relevant pricing from recent sales         
        min_price, max_price, avg_price = ebayAPIFuncLib.util_CompileSalePriceStatsOfSearchResults(historical_search_results["itemSummaries"])

        #Categorize the following motherboard historical pricing data 
        used_motherboard_historical_sale_pricing[mobo_mpn] = {"min_price": min_price,
                                                              "max_price": max_price,
                                                              "avg_price": avg_price} 

    return used_motherboard_historical_sale_pricing 


def buildMotherboardSpreadsheet(motherboard_data: dict, motherboard_used_pricing_results_by_mpn: dict, output_excel_sheet_name: str):
    '''
    #Parses scraped motherboard data and historical motherboard data and prepares it for csv format

    #COLUMN_LABELS = ["CHECKED", "BUYABLE", "EST. RETURN", "EST. RETURN %" , "BRAND", "NAME", "LISTING TITLE", "LISTING TYPE", "PRICE", "CUR AVG USED PRICE", "CUR MIN USED PRICE", "CUR MAX USED PRICE", "LISTING LINK"]

    
    #Returns a list of lists, where each list is a row and inside are the values for each column
    '''
    list_of_new_rows = [COLUMN_LABELS]
    empty_val_keyword = 0.0

    for item_row in range(0, len(motherboard_data["itemSummaries"])):
        item_listing = motherboard_data["itemSummaries"][item_row]
        new_row = [empty_val_keyword for i in range(0, len(COLUMN_LABELS))] 
        
        new_row[COLUMN_LABELS.index("CHECKED")] = "FALSE"
        new_row[COLUMN_LABELS.index("BUYABLE")] = "FALSE"
        new_row[COLUMN_LABELS.index("BRAND")] = item_listing["brand"] if "brand" in item_listing else empty_val_keyword
        new_row[COLUMN_LABELS.index("LISTING TYPE")] = ",".join(item_listing["buyingOptions"]) if "buyingOptions" in item_listing else empty_val_keyword
        
        #pricing = direct pricing + shipping costs if any 
        direct_price = float(item_listing["price"]["value"]) if "price" in item_listing else 0.0
        shipping_price = 0 
        if ("shippingOptions" in item_listing):
            shipping_price = float(item_listing["shippingOptions"][0]["shippingCost"]["value"] if "shippingCost" in item_listing["shippingOptions"][0].keys() else 0.0)
        new_row[COLUMN_LABELS.index("PRICE")] = direct_price + shipping_price
        

        #Excel hyper link formatting 
        if ("itemWebUrl" in item_listing and len(item_listing["itemWebUrl"]) > 0):
            hyper_link = item_listing["itemWebUrl"].split("?")[0] if "itemWebUrl" in item_listing else empty_val_keyword
            new_row[COLUMN_LABELS.index("LISTING LINK")] = hyper_link

        listing_title = item_listing["title"] if "title" in item_listing else "Unknown"
        new_row[COLUMN_LABELS.index("LISTING TITLE")] = "=HYPERLINK(\"" + hyper_link + "\"," + f"\"{listing_title}\")" if "itemWebUrl" in item_listing else listing_title
        

        #Formatting in reference pricing details of used motherboards of this mpn type currently for sale 
        if ("mpn" in item_listing and item_listing["mpn"] != ""):
            # try:
            new_row[COLUMN_LABELS.index("NAME")] = item_listing["mpn"]
            new_row[COLUMN_LABELS.index("CUR AVG USED PRICE")] = f"{motherboard_used_pricing_results_by_mpn[item_listing['mpn']]['avg_price']:.2f}"
            new_row[COLUMN_LABELS.index("CUR MIN USED PRICE")] = motherboard_used_pricing_results_by_mpn[item_listing["mpn"]]["min_price"]
            new_row[COLUMN_LABELS.index("CUR MAX USED PRICE")] = motherboard_used_pricing_results_by_mpn[item_listing["mpn"]]["max_price"]
            # except Exception as e:
            #     print("\nLOGGING ERROR\n")
            #     print(e)
            #     print("\n" + motherboard_used_pricing_results_by_mpn)
            #     f = open("zwamn_response_debug" + ".json", 'w')
            #     json.dump(item_listing, f, indent = 4)

            
            #calculate possible return based on minimum current going price of the mobo (used)
            new_row[COLUMN_LABELS.index("EST. RETURN")] = float(new_row[COLUMN_LABELS.index("CUR MIN USED PRICE")]) - float(new_row[COLUMN_LABELS.index("PRICE")])
            new_row[COLUMN_LABELS.index("EST. RETURN %")] = (float(new_row[COLUMN_LABELS.index('EST. RETURN')]) / float(new_row[COLUMN_LABELS.index('PRICE')])) if float(new_row[COLUMN_LABELS.index("PRICE")]) != 0.0 else 0.0

        list_of_new_rows.append(new_row)

    #If a spreadsheet already exists with the given csv file name, add its data to the current data set (skip duplicates = use the most recent acquisition of that data)
    if (output_excel_sheet_name in os.listdir() and os.path.isfile(output_excel_sheet_name)):
        mergeNewDataToExistingExcelSheet(list_of_new_rows[1:], output_excel_sheet_name) #don't include header row (first row) in list of data
    else:
        print("\nNEW Data Entries Captures...")
        excel_workbook = openpyxl.Workbook()
        active_excel_sheeet = excel_workbook.active
        for row_of_data in list_of_new_rows:
            active_excel_sheeet.append(row_of_data)
        print(f"{row_of_data[COLUMN_LABELS.index('LISTING TITLE')].split(',', maxsplit=1)[1]}; {row_of_data[COLUMN_LABELS.index('LISTING LINK')]}")
        excel_workbook.save(filename= output_excel_sheet_name)
        excel_workbook.close()



def mergeNewDataToExistingExcelSheet(new_data_rows: list[list], existing_sheet_name: str):
    excel_workbook = openpyxl.load_workbook(filename= existing_sheet_name)
    active_excel_sheet = excel_workbook.active 

#Put new data into dict because we love hashmaps! 
        #(hashmap is nice for this next part but probably not technically necessary)
    new_data_map_link_to_array_position = {}
    for i, new_data_row in enumerate(new_data_rows): 
        new_data_map_link_to_array_position[new_data_row[COLUMN_LABELS.index("LISTING LINK")]] = i  #store array index of the ebay listing 
            #using LISTING LINK as key (the link to the listing)
                # it should be unique for each item (which will help for checking duplicates in the next portion of code)
                # AND I don't expect that to change when rescanning the same listing.  


#Try to merge new rows to the existing excel sheet
        
    #First: Handle duplicates by updating existing data with new data 
        #(important that this doesn't change the order of existing rows)
    print("\nDuplicate Entries Captured...\n")
    for i, excel_row in enumerate(list(active_excel_sheet.iter_rows(min_row=2, values_only=True)), start=2): 
        #Check entry's listing link vs new data listing links
        existing_list_link = excel_row[COLUMN_LABELS.index("LISTING LINK")]

        if (existing_list_link in new_data_map_link_to_array_position):
            #update existing excel data entry

            #remove existing entry from the dictionary since we know now that it's a duplicate 
            index_of_new_data_duplicate_row = new_data_map_link_to_array_position.pop(existing_list_link) 

            print(f"{new_data_rows[index_of_new_data_duplicate_row][COLUMN_LABELS.index('LISTING TITLE')].split(',', maxsplit=1)[1]}; {new_data_rows[index_of_new_data_duplicate_row][COLUMN_LABELS.index('LISTING LINK')]}")
            
            for col in range(0, len(COLUMN_LABELS)): 
                if (col+1 == COLUMN_LABELS.index("LISTING TITLE")):
                    continue
                active_excel_sheet.cell(row = i, column = col+1).value = new_data_rows[index_of_new_data_duplicate_row][col]

    #new_data_rows without duplicates, relative to the existing worksheet removed 
    new_data_rows = [row for row in new_data_rows if row[COLUMN_LABELS.index("LISTING LINK")] in new_data_map_link_to_array_position]

    #Second: Add all remaining new data rows as true new rows
    print("\nNEW Data Entries Captures...")
    for new_data_row in new_data_rows:
        print(f"{new_data_row[COLUMN_LABELS.index('LISTING TITLE')].split(',', maxsplit=1)[1]}; {new_data_row[COLUMN_LABELS.index('LISTING LINK')]}")

# "=HYPERLINK(" + chr(ord("A") + COLUMN_LABELS.index("LISTING LINK")) + str(item_row+2)+ "," + f"\"{listing_title}\")" if "itemWebUrl" in item_listing else listing_title
        #new_listing_title_hyper_link = "=HYPERLINK(" + "[@[LISTING LINK]]," + new_data_row[COLUMN_LABELS.index('LISTING TITLE')].split(',', maxsplit=1)[1]
        #new_data_row[COLUMN_LABELS.index("LISTING TITLE")] = new_listing_title_hyper_link

#AN44, AN43, AM43, AM42
        active_excel_sheet.append(new_data_row)

#Save changes to excel sheet
    excel_workbook.save(filename= existing_sheet_name)
    excel_workbook.close()
    return


def main():
    #read command line arguments
    arguments_passed = len(sys.argv)

    #first argument will be path (and name) to a file that will become the .csv output log
    if (arguments_passed > 1):
        csv_output_file = sys.argv[1]
    else:
        csv_output_file = "motherboard_data" + str(datetime.date(datetime.now())) + ".csv"


    compileMotherboardData(csv_output_file)
    return

main()